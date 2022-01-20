from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime


def execution_time(func):
    def wrapper(*args, **kwargs):
        initial_time = datetime.now()
        func(*args, **kwargs)
        final_time = datetime.now()
        time_elapsed = final_time - initial_time
        print(f'Pasaron {time_elapsed.total_seconds()} segundos')

    return wrapper


print("Esto es la rama v_2.0")

"""from colorama import init
from termcolor import colored"""

"""Cosas por mejorar:
1) guardar la contraseña en un lugar seguro
2) Poner el logo en el correo electrónico
3) correspondencia con diferentes archivos
5) evaluar los excel para automatizar los rangos
6) arreglar lo de los inmuebles con más correos"""

"""print(ws['A2'].value)"""


def obtener_valores() -> list:
    wb = load_workbook('Formato correos nuevo.xlsx')
    ws = wb.active
    lista_completa = []
    filas_total = 3  # número de la última fila
    columnas_total = 9
    for row in range(2, filas_total + 1):  # Se tienen que cambiar
        laux = []
        for col in range(1, columnas_total + 1):  # Se tienen que cambiar
            chr = get_column_letter(col)
            laux.append(ws[chr + str(row)].value)
        lista_completa.append(laux)
    return lista_completa


def limpiar_datos(lista: list) -> tuple:
    lclean = []
    for i in lista:
        laux = []
        for j in i:
            ty = type(j)
            if ty == int or ty == float:
                laux.append(round(j))
            elif j is None or j == 0:
                laux.append(0)
            elif ty == str:
                """if ';' in j:
                    laux.append(j.split(';'))
                else:"""
                laux.append(j)
        lclean.append(laux)
    return tuple(lclean)


@execution_time
def enviar_mensaje(lista: tuple) -> None:
    count = 0
    for row in lista:
        inmueble = row[0]
        capital_inicial = row[1]
        honorarios = row[2]
        costas = row[3]
        total = row[4]
        correos = row[5].strip().lower()  # Repetir en caso de dos o más correos!
        asunto = row[6].strip()
        tipo_cobro = row[7].strip()
        tipo_AP = row[8].strip()

        fecha = '31 de diciembre de 2021'

        mensaje = MIMEMultipart()
        mensaje['Subject'] = f'COBRO {tipo_cobro.upper()} - {asunto} - INMUEBLE {inmueble}'

        cuerpo = f"""\
        <!DOCTYPE html>
        <html>
          <head></head>
          <body>
              <header>
                  <a href="">
                      <img src="C:Users\AMD\Documents\CARVEL\Python\Correos\logo.png" alt="logo" width="100" height="100">
                  <h3><b><ins>RECORDATORIO COBRO {tipo_cobro.upper()}</h3><br><br>
              </header> 
              <main>
                <section>
                  <p>Apreciado(a) Señor (a),<br><br><br></p>
                  <p>La obligación que usted tiene como propietario/tenedor del Inmueble de la referencia, se encuentra actualmente en cobro {tipo_AP.upper()} bajo la supervisión de CARVEL SOLUCIONES JURÍDICAS Y ADMINISTRATIVAS S.A.S., lo anterior debido al incumplimiento en sus pagos de las cuotas de administración que a <b>{fecha}</b> se discriminan a continuación:</p><br><br>
                  <p>Saldo a Capital: <b>${capital_inicial + costas:,.0f}</p>
                  <p>Honorarios Cobro {tipo_cobro}: <b>${honorarios:,.0f}</p>
                  <p>Total Deuda: <b>${total:,.0f}</p><br><br>
                  <p>Solicitamos cancelar el valor adeudado y si ya realizó el pago, le pedimos que por favor se comunique con nosotros para verificar con los respectivos soportes en la administración.</p>
                  <p>Es necesario que ponga al día la obligación, debido a que si usted hace abonos parciales estos deben estar autorizados previamente, ya que es usted acreedor al pago de honorarios sobre todos los abonos que efectúe.</p>
                  <p><b>AÚN PUEDE EVITAR ESTA SITUACIÓN, COMUNÍQUESE DE INMEDIATO</b> y así impedir más cargos, con el fin de efectuar el pago y/o llegar a un acuerdo.
                  Estaremos atentos a cualquier inquietud en nuestras líneas <b>3107619044</b> o en el correo electrónico
                  <a href = "mailto: carvel.soluciones@gmail.com">carvel.soluciones@gmail.com.</a><br><br><br>
                  <p>
                  Cordialmente,<br><br><br>
                  <b>Carolina Velásquez</b><br>
                  <b>Departamento Jurídico</b> 
                  </p>
                </section>
              </main> 
          </body>
        </html>
        """
        mensaje.attach(MIMEText(cuerpo, 'html', 'utf-8'))
        texto = mensaje.as_string()
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login('carvel.soluciones@gmail.com', 'Chico2021*')
        servidor.sendmail('carvel.soluciones@gmail.com', correos, texto)
        servidor.quit()

        count += 1
        print(f'{count}. Envio exitoso')

    print(f'\nFin del proceso\nTotal correos enviados: {count}\n')


# ------------------------------------------------------

def main():
    ov = obtener_valores()
    ld = limpiar_datos(ov)
    enviar_mensaje(ld)


if __name__ == '__main__':
    main()
